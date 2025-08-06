from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, Count, F, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.timezone import localtime
from django.views import View
from django.views.generic import (
    ListView,
    DetailView,
    CreateView,
    UpdateView,
    DeleteView,
)

import openpyxl

from crm_entities.models import Account, Contact, Lead
from sales_pipeline.models import Deal
from sales_territories.models import Territory
from users.models import CustomUser

from .filters import TaskFilter, CallFilter, MeetingFilter
from .forms import TaskForm, CallForm, MeetingForm
from .models import Task, Call, Meeting, TaskStatusChoices, CallMeetingStatusChoices, TaskPriorityChoices


class BaseActivityView(LoginRequiredMixin):
    """Mixin to require login and provide role-based filtering."""

    def _filter_queryset_by_role(self, user, queryset):
        if not hasattr(self, 'model'):
            return queryset.none()

        if user.is_admin_role:
            return queryset
        elif user.is_manager_role:
            try:
                managed_territories = user.managed_territories.all()
                team_members = CustomUser.objects.filter(
                    territory__in=managed_territories,
                    role=CustomUser.Roles.SALES,
                ).exclude(pk=user.pk)
                base_q = (
                    Q(assigned_to=user) |
                    Q(created_by=user) |
                    Q(assigned_to__in=team_members) |
                    Q(created_by__in=team_members)
                )
                return queryset.filter(base_q).distinct()
            except Exception as e:
                print(f"Error applying manager role filter in activities for {self.model.__name__}: {e}")
                return queryset.filter(
                    Q(assigned_to=user) | Q(created_by=user)
                ).distinct()
        elif user.is_sales_role:
            return queryset.filter(
                Q(assigned_to=user) | Q(created_by=user)
            ).distinct()
        return queryset.none()


class TaskListView(BaseActivityView, ListView):
    model = Task
    context_object_name = 'tasks'
    template_name = 'activities/task_list.html'
    paginate_by = 20
    sort_by_applied = 'due_date'
    direction_applied = 'asc'

    def get_queryset(self):
        user = self.request.user
        base_queryset = Task.objects.all()
        queryset = self._filter_queryset_by_role(user, base_queryset)

        filter_params = self.request.GET.copy()
        for param in ['sort', 'dir', 'page']:
            filter_params.pop(param, None)
        self.filterset = TaskFilter(filter_params, queryset=queryset)
        queryset = self.filterset.qs

        sort_by_param = self.request.GET.get('sort', 'due_date')
        direction_param = self.request.GET.get('dir', 'asc')
        valid_sort_fields = [
            'subject',
            'status',
            'priority',
            'due_date',
            'assigned_to__username',
            'updated_at',
        ]
        sort_by_validated = sort_by_param.lstrip('-')
        if sort_by_validated not in valid_sort_fields:
            sort_by_validated = 'due_date'
            direction_validated = 'asc'
        else:
            direction_validated = 'desc' if direction_param == 'desc' else 'asc'
        self.sort_by_applied = sort_by_validated
        self.direction_applied = direction_validated

        sort_by_final = (
            f'-{self.sort_by_applied}' if self.direction_applied == 'desc'
            else self.sort_by_applied
        )
        return queryset.order_by(sort_by_final).select_related(
            'assigned_to',
            'created_by',
            'related_to_account',
            'related_to_contact',
            'related_to_lead',
            'related_to_deal',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filterset'] = self.filterset
        context['sort_by'] = self.sort_by_applied
        context['direction'] = self.direction_applied
        context['opposite_direction'] = (
            'desc' if self.direction_applied == 'asc' else 'asc'
        )
        context['title'] = "Tasks"
        query_params = self.request.GET.copy()
        for param in ['sort', 'dir', 'page']:
            query_params.pop(param, None)
        context['current_filters_encoded'] = query_params.urlencode()
        return context


class TaskDetailView(BaseActivityView, DetailView):
    model = Task
    context_object_name = 'task'
    template_name = 'activities/task_detail.html'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related(
            'assigned_to',
            'created_by',
            'related_to_account',
            'related_to_contact',
            'related_to_lead',
            'related_to_deal',
        )
        return self._filter_queryset_by_role(user, queryset)


class TaskCreateView(BaseActivityView, CreateView):
    model = Task
    form_class = TaskForm
    template_name = 'activities/task_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Create New Task'
        return context

    def get_initial(self):
        initial = super().get_initial()
        if self.request.user.is_authenticated:
            initial['assigned_to'] = self.request.user

        account_pk = self.request.GET.get('account')
        contact_pk = self.request.GET.get('contact')
        lead_pk = self.request.GET.get('lead')
        deal_pk = self.request.GET.get('deal')

        if account_pk:
            try:
                initial['related_to_account'] = Account.objects.get(
                    pk=account_pk
                )
            except Account.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Account (ID:{account_pk}) specified."
                )
        elif contact_pk:
            try:
                initial['related_to_contact'] = Contact.objects.get(
                    pk=contact_pk
                )
            except Contact.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Contact (ID:{contact_pk}) specified."
                )
        elif lead_pk:
            try:
                initial['related_to_lead'] = Lead.objects.get(pk=lead_pk)
            except Lead.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Lead (ID:{lead_pk}) specified."
                )
        elif deal_pk:
            try:
                initial['related_to_deal'] = Deal.objects.get(pk=deal_pk)
            except Deal.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Deal (ID:{deal_pk}) specified."
                )
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if hasattr(form.instance, 'assigned_to') and not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if self.request.GET.get('account'):
            return reverse(
                'crm_entities:account-detail',
                kwargs={'pk': self.request.GET.get('account')}
            )
        elif self.request.GET.get('contact'):
            return reverse(
                'crm_entities:contact-detail',
                kwargs={'pk': self.request.GET.get('contact')}
            )
        elif self.request.GET.get('lead'):
            return reverse(
                'crm_entities:lead-detail',
                kwargs={'pk': self.request.GET.get('lead')}
            )
        elif self.request.GET.get('deal'):
            return reverse(
                'sales_pipeline:deal-detail',
                kwargs={'pk': self.request.GET.get('deal')}
            )
        return reverse_lazy('activities:task-list')


class TaskUpdateView(BaseActivityView, UpdateView):
    model = Task
    form_class = TaskForm
    template_name = 'activities/task_form.html'
    success_url = reverse_lazy('activities:task-list')

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Update Task: {self.object.subject}'
        return context


class TaskDeleteView(BaseActivityView, DeleteView):
    model = Task
    template_name = 'activities/task_confirm_delete.html'
    success_url = reverse_lazy('activities:task-list')
    context_object_name = 'task'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)


class CallListView(BaseActivityView, ListView):
    model = Call
    context_object_name = 'calls'
    template_name = 'activities/call_list.html'
    paginate_by = 20
    sort_by_applied = '-call_time'
    direction_applied = 'desc'

    def get_queryset(self):
        user = self.request.user
        base_queryset = Call.objects.all()
        queryset = self._filter_queryset_by_role(user, base_queryset)
        filter_params = self.request.GET.copy()
        for param in ['sort', 'dir', 'page']:
            filter_params.pop(param, None)
        self.filterset = CallFilter(filter_params, queryset=queryset)
        queryset = self.filterset.qs
        sort_by_param = self.request.GET.get('sort', '-call_time')
        direction_param = self.request.GET.get('dir', 'desc')
        valid_sort_fields = [
            'subject',
            'call_time',
            'direction',
            'status',
            'assigned_to__username',
            'updated_at',
        ]
        sort_by_validated = sort_by_param.lstrip('-')
        if sort_by_validated not in valid_sort_fields:
            sort_by_validated = 'call_time'
            direction_validated = 'desc'
        else:
            direction_validated = 'desc' if direction_param == 'desc' else 'asc'
        self.sort_by_applied = sort_by_validated
        self.direction_applied = direction_validated
        sort_by_final = (
            f'-{self.sort_by_applied}' if self.direction_applied == 'desc'
            else self.sort_by_applied
        )
        return queryset.order_by(sort_by_final).select_related(
            'assigned_to',
            'created_by',
            'related_to_account',
            'related_to_contact',
            'related_to_lead',
            'related_to_deal',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filterset'] = self.filterset
        context['sort_by'] = self.sort_by_applied
        context['direction'] = self.direction_applied
        context['opposite_direction'] = (
            'desc' if self.direction_applied == 'asc' else 'asc'
        )
        context['title'] = "Calls"
        query_params = self.request.GET.copy()
        for param in ['sort', 'dir', 'page']:
            query_params.pop(param, None)
        context['current_filters_encoded'] = query_params.urlencode()
        return context


class CallDetailView(BaseActivityView, DetailView):
    model = Call
    context_object_name = 'call'
    template_name = 'activities/call_detail.html'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related(
            'assigned_to',
            'created_by',
            'related_to_account',
            'related_to_contact',
            'related_to_lead',
            'related_to_deal',
        )
        return self._filter_queryset_by_role(user, queryset)


class CallCreateView(BaseActivityView, CreateView):
    model = Call
    form_class = CallForm
    template_name = 'activities/call_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Log New Call'
        return context

    def get_initial(self):
        initial = super().get_initial()
        if self.request.user.is_authenticated:
            initial['assigned_to'] = self.request.user

        account_pk = self.request.GET.get('account')
        contact_pk = self.request.GET.get('contact')
        lead_pk = self.request.GET.get('lead')
        deal_pk = self.request.GET.get('deal')

        if account_pk:
            try:
                initial['related_to_account'] = Account.objects.get(
                    pk=account_pk
                )
            except Account.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Account (ID:{account_pk}) specified."
                )
        elif contact_pk:
            try:
                initial['related_to_contact'] = Contact.objects.get(
                    pk=contact_pk
                )
            except Contact.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Contact (ID:{contact_pk}) specified."
                )
        elif lead_pk:
            try:
                initial['related_to_lead'] = Lead.objects.get(pk=lead_pk)
            except Lead.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Lead (ID:{lead_pk}) specified."
                )
        elif deal_pk:
            try:
                initial['related_to_deal'] = Deal.objects.get(pk=deal_pk)
            except Deal.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Deal (ID:{deal_pk}) specified."
                )
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if hasattr(form.instance, 'assigned_to') and not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if self.request.GET.get('account'):
            return reverse(
                'crm_entities:account-detail',
                kwargs={'pk': self.request.GET.get('account')}
            )
        elif self.request.GET.get('contact'):
            return reverse(
                'crm_entities:contact-detail',
                kwargs={'pk': self.request.GET.get('contact')}
            )
        elif self.request.GET.get('lead'):
            return reverse(
                'crm_entities:lead-detail',
                kwargs={'pk': self.request.GET.get('lead')}
            )
        elif self.request.GET.get('deal'):
            return reverse(
                'sales_pipeline:deal-detail',
                kwargs={'pk': self.request.GET.get('deal')}
            )
        return reverse_lazy('activities:call-list')


class CallUpdateView(BaseActivityView, UpdateView):
    model = Call
    form_class = CallForm
    template_name = 'activities/call_form.html'
    success_url = reverse_lazy('activities:call-list')

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Update Call: {self.object.subject}'
        return context


class CallDeleteView(BaseActivityView, DeleteView):
    model = Call
    template_name = 'activities/call_confirm_delete.html'
    success_url = reverse_lazy('activities:call-list')
    context_object_name = 'call'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)


class MeetingListView(BaseActivityView, ListView):
    model = Meeting
    context_object_name = 'meetings'
    template_name = 'activities/meeting_list.html'
    paginate_by = 20
    sort_by_applied = '-start_time'
    direction_applied = 'desc'

    def get_queryset(self):
        user = self.request.user
        base_queryset = Meeting.objects.all()
        queryset = self._filter_queryset_by_role(user, base_queryset)
        filter_params = self.request.GET.copy()
        for param in ['sort', 'dir', 'page']:
            filter_params.pop(param, None)
        self.filterset = MeetingFilter(filter_params, queryset=queryset)
        queryset = self.filterset.qs
        sort_by_param = self.request.GET.get('sort', '-start_time')
        direction_param = self.request.GET.get('dir', 'desc')
        valid_sort_fields = [
            'subject',
            'start_time',
            'location',
            'status',
            'assigned_to__username',
            'updated_at',
        ]
        sort_by_validated = sort_by_param.lstrip('-')
        if sort_by_validated not in valid_sort_fields:
            sort_by_validated = 'start_time'
            direction_validated = 'desc'
        else:
            direction_validated = 'desc' if direction_param == 'desc' else 'asc'
        self.sort_by_applied = sort_by_validated
        self.direction_applied = direction_validated
        sort_by_final = (
            f'-{self.sort_by_applied}' if self.direction_applied == 'desc'
            else self.sort_by_applied
        )
        return queryset.order_by(sort_by_final).select_related(
            'assigned_to',
            'created_by',
            'related_to_account',
            'related_to_contact',
            'related_to_lead',
            'related_to_deal',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filterset'] = self.filterset
        context['sort_by'] = self.sort_by_applied
        context['direction'] = self.direction_applied
        context['opposite_direction'] = (
            'desc' if self.direction_applied == 'asc' else 'asc'
        )
        context['title'] = "Meetings"
        query_params = self.request.GET.copy()
        for param in ['sort', 'dir', 'page']:
            query_params.pop(param, None)
        context['current_filters_encoded'] = query_params.urlencode()
        return context


class MeetingDetailView(BaseActivityView, DetailView):
    model = Meeting
    context_object_name = 'meeting'
    template_name = 'activities/meeting_detail.html'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related(
            'assigned_to',
            'created_by',
            'related_to_account',
            'related_to_contact',
            'related_to_lead',
            'related_to_deal',
        )
        return self._filter_queryset_by_role(user, queryset)


class MeetingCreateView(BaseActivityView, CreateView):
    model = Meeting
    form_class = MeetingForm
    template_name = 'activities/meeting_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Schedule New Meeting'
        return context

    def get_initial(self):
        initial = super().get_initial()
        now = timezone.now().replace(second=0, microsecond=0)
        initial['start_time'] = now
        initial['end_time'] = now + timedelta(minutes=30)

        if self.request.user.is_authenticated:
            initial['assigned_to'] = self.request.user

        account_pk = self.request.GET.get('account')
        contact_pk = self.request.GET.get('contact')
        lead_pk = self.request.GET.get('lead')
        deal_pk = self.request.GET.get('deal')

        if account_pk:
            try:
                initial['related_to_account'] = Account.objects.get(
                    pk=account_pk
                )
            except Account.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Account (ID:{account_pk}) specified."
                )
        elif contact_pk:
            try:
                initial['related_to_contact'] = Contact.objects.get(
                    pk=contact_pk
                )
            except Contact.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Contact (ID:{contact_pk}) specified."
                )
        elif lead_pk:
            try:
                initial['related_to_lead'] = Lead.objects.get(pk=lead_pk)
            except Lead.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Lead (ID:{lead_pk}) specified."
                )
        elif deal_pk:
            try:
                initial['related_to_deal'] = Deal.objects.get(pk=deal_pk)
            except Deal.DoesNotExist:
                messages.error(
                    self.request,
                    f"Invalid Deal (ID:{deal_pk}) specified."
                )
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if hasattr(form.instance, 'assigned_to') and not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if self.request.GET.get('account'):
            return reverse(
                'crm_entities:account-detail',
                kwargs={'pk': self.request.GET.get('account')}
            )
        elif self.request.GET.get('contact'):
            return reverse(
                'crm_entities:contact-detail',
                kwargs={'pk': self.request.GET.get('contact')}
            )
        elif self.request.GET.get('lead'):
            return reverse(
                'crm_entities:lead-detail',
                kwargs={'pk': self.request.GET.get('lead')}
            )
        elif self.request.GET.get('deal'):
            return reverse(
                'sales_pipeline:deal-detail',
                kwargs={'pk': self.request.GET.get('deal')}
            )
        return reverse_lazy('activities:meeting-list')


class MeetingUpdateView(BaseActivityView, UpdateView):
    model = Meeting
    form_class = MeetingForm
    template_name = 'activities/meeting_form.html'
    success_url = reverse_lazy('activities:meeting-list')

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Update Meeting: {self.object.subject}'
        return context


class MeetingDeleteView(BaseActivityView, DeleteView):
    model = Meeting
    template_name = 'activities/meeting_confirm_delete.html'
    success_url = reverse_lazy('activities:meeting-list')
    context_object_name = 'meeting'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)


@login_required
def task_export_view(request):
    if not request.user.is_admin_role:
        return HttpResponseForbidden("Permission Denied.")

    user = request.user
    base_queryset = Task.objects.all()
    filter_params = request.GET.copy()
    for param in ['sort', 'dir', 'page']:
        filter_params.pop(param, None)
    filterset = TaskFilter(filter_params, queryset=base_queryset)
    queryset = filterset.qs
    sort_by_param = request.GET.get('sort', 'due_date')
    direction_param = request.GET.get('dir', 'asc')
    valid_sort_fields = [
        'subject',
        'status',
        'priority',
        'due_date',
        'assigned_to__username',
        'updated_at',
    ]
    sort_by_validated = sort_by_param.lstrip('-')
    if sort_by_validated not in valid_sort_fields:
        sort_by_validated = 'due_date'
        direction_validated = 'asc'
    else:
        direction_validated = 'desc' if direction_param == 'desc' else 'asc'
    sort_by_final = (
        f'-{sort_by_validated}' if direction_validated == 'desc'
        else sort_by_validated
    )
    queryset = queryset.order_by(sort_by_final).select_related(
        'assigned_to',
        'created_by',
        'related_to_account',
        'related_to_contact',
        'related_to_lead',
        'related_to_deal',
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    headers = [
        "ID",
        "Subject",
        "Status",
        "Priority",
        "Due Date",
        "Related Account",
        "Related Contact",
        "Related Lead",
        "Related Deal",
        "Description",
        "Assigned To",
        "Created By",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    for task in queryset:
        assigned_to_name = (
            task.assigned_to.get_full_name() or
            task.assigned_to.username if task.assigned_to else ""
        )
        created_by_name = (
            task.created_by.get_full_name() or
            task.created_by.username if task.created_by else ""
        )
        created_at_formatted = (
            localtime(task.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if task.created_at else ""
        )
        updated_at_formatted = (
            localtime(task.updated_at).strftime('%Y-%m-%d %H:%M:%S')
            if task.updated_at else ""
        )
        related_account = (
            task.related_to_account.name if task.related_to_account else ""
        )
        related_contact = (
            task.related_to_contact.full_name
            if task.related_to_contact else ""
        )
        related_lead = (
            task.related_to_lead.full_name if task.related_to_lead else ""
        )
        related_deal = str(task.related_to_deal) if task.related_to_deal else ""
        row = [
            task.pk,
            task.subject or "",
            task.get_status_display(),
            task.get_priority_display(),
            task.due_date,
            related_account,
            related_contact,
            related_lead,
            related_deal,
            task.description or "",
            assigned_to_name,
            created_by_name,
            created_at_formatted,
            updated_at_formatted,
        ]
        ws.append(row)

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = 'attachment; filename="tasks_export.xlsx"'
    wb.save(response)
    return response


@login_required
def call_export_view(request):
    if not request.user.is_admin_role:
        return HttpResponseForbidden("Permission Denied.")

    user = request.user
    base_queryset = Call.objects.all()
    filter_params = request.GET.copy()
    for param in ['sort', 'dir', 'page']:
        filter_params.pop(param, None)
    filterset = CallFilter(filter_params, queryset=base_queryset)
    queryset = filterset.qs
    sort_by_param = request.GET.get('sort', '-call_time')
    direction_param = request.GET.get('dir', 'desc')
    valid_sort_fields = [
        'subject',
        'call_time',
        'direction',
        'status',
        'assigned_to__username',
        'updated_at',
    ]
    sort_by_validated = sort_by_param.lstrip('-')
    if sort_by_validated not in valid_sort_fields:
        sort_by_validated = 'call_time'
        direction_validated = 'desc'
    else:
        direction_validated = 'desc' if direction_param == 'desc' else 'asc'
    sort_by_final = (
        f'-{sort_by_validated}' if direction_validated == 'desc'
        else sort_by_validated
    )
    queryset = queryset.order_by(sort_by_final).select_related(
        'assigned_to',
        'created_by',
        'related_to_account',
        'related_to_contact',
        'related_to_lead',
        'related_to_deal',
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calls"
    headers = [
        "ID",
        "Subject",
        "Call Time",
        "Duration (Min)",
        "Direction",
        "Status",
        "Related Account",
        "Related Contact",
        "Related Lead",
        "Related Deal",
        "Notes",
        "Assigned To",
        "Created By",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    for call in queryset:
        assigned_to_name = (
            call.assigned_to.get_full_name() or
            call.assigned_to.username if call.assigned_to else ""
        )
        created_by_name = (
            call.created_by.get_full_name() or
            call.created_by.username if call.created_by else ""
        )
        created_at_formatted = (
            localtime(call.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if call.created_at else ""
        )
        updated_at_formatted = (
            localtime(call.updated_at).strftime('%Y-%m-%d %H:%M:%S')
            if call.updated_at else ""
        )
        call_time_formatted = (
            localtime(call.call_time).strftime('%Y-%m-%d %H:%M:%S')
            if call.call_time else ""
        )
        related_account = (
            call.related_to_account.name if call.related_to_account else ""
        )
        related_contact = (
            call.related_to_contact.full_name
            if call.related_to_contact else ""
        )
        related_lead = (
            call.related_to_lead.full_name if call.related_to_lead else ""
        )
        related_deal = str(call.related_to_deal) if call.related_to_deal else ""
        row = [
            call.pk,
            call.subject or "",
            call_time_formatted,
            call.duration_minutes,
            call.get_direction_display(),
            call.get_status_display(),
            related_account,
            related_contact,
            related_lead,
            related_deal,
            call.notes or "",
            assigned_to_name,
            created_by_name,
            created_at_formatted,
            updated_at_formatted,
        ]
        ws.append(row)

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = 'attachment; filename="calls_export.xlsx"'
    wb.save(response)
    return response


@login_required
def meeting_export_view(request):
    if not request.user.is_admin_role:
        return HttpResponseForbidden("Permission Denied.")

    user = request.user
    base_queryset = Meeting.objects.all()
    filter_params = request.GET.copy()
    for param in ['sort', 'dir', 'page']:
        filter_params.pop(param, None)
    filterset = MeetingFilter(filter_params, queryset=base_queryset)
    queryset = filterset.qs
    sort_by_param = request.GET.get('sort', '-start_time')
    direction_param = request.GET.get('dir', 'desc')
    valid_sort_fields = [
        'subject',
        'start_time',
        'location',
        'status',
        'assigned_to__username',
        'updated_at',
    ]
    sort_by_validated = sort_by_param.lstrip('-')
    if sort_by_validated not in valid_sort_fields:
        sort_by_validated = 'start_time'
        direction_validated = 'desc'
    else:
        direction_validated = 'desc' if direction_param == 'desc' else 'asc'
    sort_by_final = (
        f'-{sort_by_validated}' if direction_validated == 'desc'
        else sort_by_validated
    )
    queryset = queryset.order_by(sort_by_final).select_related(
        'assigned_to',
        'created_by',
        'related_to_account',
        'related_to_contact',
        'related_to_lead',
        'related_to_deal',
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meetings"
    headers = [
        "ID",
        "Subject",
        "Start Time",
        "End Time",
        "Location",
        "Status",
        "Related Account",
        "Related Contact",
        "Related Lead",
        "Related Deal",
        "Description",
        "Assigned To",
        "Created By",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    for meeting in queryset:
        assigned_to_name = (
            meeting.assigned_to.get_full_name() or
            meeting.assigned_to.username if meeting.assigned_to else ""
        )
        created_by_name = (
            meeting.created_by.get_full_name() or
            meeting.created_by.username if meeting.created_by else ""
        )
        created_at_formatted = (
            localtime(meeting.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if meeting.created_at else ""
        )
        updated_at_formatted = (
            localtime(meeting.updated_at).strftime('%Y-%m-%d %H:%M:%S')
            if meeting.updated_at else ""
        )
        start_time_formatted = (
            localtime(meeting.start_time).strftime('%Y-%m-%d %H:%M:%S')
            if meeting.start_time else ""
        )
        end_time_formatted = (
            localtime(meeting.end_time).strftime('%Y-%m-%d %H:%M:%S')
            if meeting.end_time else ""
        )
        related_account = (
            meeting.related_to_account.name
            if meeting.related_to_account else ""
        )
        related_contact = (
            meeting.related_to_contact.full_name
            if meeting.related_to_contact else ""
        )
        related_lead = (
            meeting.related_to_lead.full_name if meeting.related_to_lead else ""
        )
        related_deal = (
            str(meeting.related_to_deal) if meeting.related_to_deal else ""
        )
        row = [
            meeting.pk,
            meeting.subject or "",
            start_time_formatted,
            end_time_formatted,
            meeting.location or "",
            meeting.get_status_display(),
            related_account,
            related_contact,
            related_lead,
            related_deal,
            meeting.description or "",
            assigned_to_name,
            created_by_name,
            created_at_formatted,
            updated_at_formatted,
        ]
        ws.append(row)

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = 'attachment; filename="meetings_export.xlsx"'
    wb.save(response)
    return response