from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.timezone import localtime
from django.views import View
from django.views.generic import (
    ListView,
    DetailView,
    CreateView,
    UpdateView,
    DeleteView,
)
from datetime import timedelta
import openpyxl

from dal import autocomplete

from .filters import AccountFilter, ContactFilter, LeadFilter
from .forms import AccountForm, ContactForm, LeadForm
from .models import Account, Contact, Lead
from sales_pipeline.models import Deal
from sales_territories.models import Territory
from users.models import CustomUser


class BaseCrmView(LoginRequiredMixin):
    """Basic Mixin to require login and contain role filtering helper"""

    def _filter_queryset_by_role(self, user, queryset):
        if not hasattr(self, 'model'):
            return queryset.none()

        if user.is_admin_role:
            return queryset
        elif user.is_manager_role:
            try:
                managed_territories = user.managed_territories.all()
                team_members = CustomUser.objects.filter(
                    territory__in=managed_territories,
                    role=CustomUser.Roles.SALES,
                ).exclude(pk=user.pk)

                base_q = (
                    Q(assigned_to=user) |
                    Q(created_by=user) |
                    Q(assigned_to__in=team_members) |
                    Q(created_by__in=team_members)
                )

                if hasattr(self.model, 'territory') and hasattr(Territory, 'objects'):
                    base_q |= Q(territory__in=managed_territories)
                elif hasattr(self.model, 'account') and hasattr(Account, 'territory'):
                    base_q |= Q(account__territory__in=managed_territories)

                return queryset.filter(base_q).distinct()
            except Exception as e:
                print(f"Error applying manager role filter for {self.model.__name__}: {e}")
                return queryset.filter(
                    Q(assigned_to=user) | Q(created_by=user)
                ).distinct()
        elif user.is_sales_role:
            return queryset.filter(
                Q(assigned_to=user) | Q(created_by=user)
            ).distinct()
        else:
            return queryset.none()


class AccountListView(BaseCrmView, ListView):
    model = Account
    context_object_name = 'accounts'
    template_name = 'crm_entities/account_list.html'
    paginate_by = 15
    sort_by_applied = 'name'
    direction_applied = 'asc'

    def get_queryset(self):
        user = self.request.user
        base_queryset = Account.objects.all()
        queryset = self._filter_queryset_by_role(user, base_queryset)

        filter_params = self.request.GET.copy()
        filter_params.pop('sort', None)
        filter_params.pop('dir', None)
        filter_params.pop('page', None)
        self.filterset = AccountFilter(filter_params, queryset=queryset)
        queryset = self.filterset.qs

        sort_by_param = self.request.GET.get('sort', 'name')
        direction_param = self.request.GET.get('dir', 'asc')
        valid_sort_fields = [
            'name',
            'status',
            'territory__name',
            'assigned_to__username',
            'updated_at',
        ]
        sort_by_validated = sort_by_param.lstrip('-')
        if sort_by_validated not in valid_sort_fields:
            sort_by_validated = 'name'
            direction_validated = 'asc'
        else:
            direction_validated = 'desc' if direction_param == 'desc' else 'asc'
        self.sort_by_applied = sort_by_validated
        self.direction_applied = direction_validated
        sort_by_final = (
            f'-{self.sort_by_applied}' if self.direction_applied == 'desc'
            else self.sort_by_applied
        )

        queryset = queryset.order_by(sort_by_final).select_related(
            'territory',
            'assigned_to',
        )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filterset'] = self.filterset
        context['sort_by'] = self.sort_by_applied
        context['direction'] = self.direction_applied
        context['opposite_direction'] = (
            'desc' if self.direction_applied == 'asc' else 'asc'
        )
        query_params = self.request.GET.copy()
        query_params.pop('sort', None)
        query_params.pop('dir', None)
        query_params.pop('page', None)
        context['current_filters_encoded'] = query_params.urlencode()
        return context


class AccountDetailView(BaseCrmView, DetailView):
    model = Account
    context_object_name = 'account'
    template_name = 'crm_entities/account_detail.html'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related(
            'territory',
            'assigned_to',
            'created_by',
        )
        return self._filter_queryset_by_role(user, queryset)


class AccountCreateView(BaseCrmView, CreateView):
    model = Account
    form_class = AccountForm
    template_name = 'crm_entities/account_form.html'
    success_url = reverse_lazy('crm_entities:account-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Create New Account'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if hasattr(form.instance, 'assigned_to') and not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)


class AccountUpdateView(BaseCrmView, UpdateView):
    model = Account
    form_class = AccountForm
    template_name = 'crm_entities/account_form.html'
    success_url = reverse_lazy('crm_entities:account-list')

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Update Account: {self.object.name}'
        return context


class AccountDeleteView(PermissionRequiredMixin, BaseCrmView, DeleteView):
    model = Account
    permission_required = 'crm_entities.delete_account' # Permission codename format: app_label.verb_modelname
    template_name = 'crm_entities/account_confirm_delete.html'
    success_url = reverse_lazy('crm_entities:account-list')
    context_object_name = 'account'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)


class ContactListView(BaseCrmView, ListView):
    model = Contact
    context_object_name = 'contacts'
    template_name = 'crm_entities/contact_list.html'
    paginate_by = 15
    sort_by_applied = 'last_name'
    direction_applied = 'asc'

    def get_queryset(self):
        user = self.request.user
        base_queryset = Contact.objects.all()
        queryset = self._filter_queryset_by_role(user, base_queryset)
        filter_params = self.request.GET.copy()
        filter_params.pop('sort', None)
        filter_params.pop('dir', None)
        filter_params.pop('page', None)
        self.filterset = ContactFilter(filter_params, queryset=queryset)
        queryset = self.filterset.qs

        sort_by_param = self.request.GET.get('sort', 'last_name')
        direction_param = self.request.GET.get('dir', 'asc')
        valid_sort_fields = [
            'last_name',
            'first_name',
            'account__name',
            'title',
            'department',
            'email',
            'assigned_to__username',
            'updated_at',
        ]
        sort_by_validated = sort_by_param.lstrip('-')
        if sort_by_validated not in valid_sort_fields:
            sort_by_validated = 'last_name'
            direction_validated = 'asc'
        else:
            direction_validated = 'desc' if direction_param == 'desc' else 'asc'
        self.sort_by_applied = sort_by_validated
        self.direction_applied = direction_validated
        sort_by_final = (
            f'-{self.sort_by_applied}' if self.direction_applied == 'desc'
            else self.sort_by_applied
        )
        queryset = queryset.order_by(sort_by_final).select_related(
            'account',
            'assigned_to',
        )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filterset'] = self.filterset
        context['sort_by'] = self.sort_by_applied
        context['direction'] = self.direction_applied
        context['opposite_direction'] = (
            'desc' if self.direction_applied == 'asc' else 'asc'
        )
        query_params = self.request.GET.copy()
        query_params.pop('sort', None)
        query_params.pop('dir', None)
        query_params.pop('page', None)
        context['current_filters_encoded'] = query_params.urlencode()
        return context


class ContactDetailView(BaseCrmView, DetailView):
    model = Contact
    context_object_name = 'contact'
    template_name = 'crm_entities/contact_detail.html'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related(
            'account',
            'assigned_to',
            'created_by',
        )
        return self._filter_queryset_by_role(user, queryset)


class ContactCreateView(BaseCrmView, CreateView):
    model = Contact
    form_class = ContactForm
    template_name = 'crm_entities/contact_form.html'
    success_url = reverse_lazy('crm_entities:contact-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Create New Contact'
        return context

    def get_initial(self):
        initial = super().get_initial()
        account_pk = self.request.GET.get('account')
        if account_pk:
            try:
                initial['account'] = Account.objects.get(pk=account_pk)
            except Account.DoesNotExist:
                messages.error(self.request, "Invalid Account specified.")
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if hasattr(form.instance, 'assigned_to') and not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)


class ContactUpdateView(BaseCrmView, UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = 'crm_entities/contact_form.html'
    success_url = reverse_lazy('crm_entities:contact-list')

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Update Contact: {self.object.full_name}'
        return context


class ContactDeleteView(BaseCrmView, DeleteView):
    model = Contact
    template_name = 'crm_entities/contact_confirm_delete.html'
    success_url = reverse_lazy('crm_entities:contact-list')
    context_object_name = 'contact'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)


class LeadListView(BaseCrmView, ListView):
    model = Lead
    context_object_name = 'leads'
    template_name = 'crm_entities/lead_list.html'
    paginate_by = 15
    sort_by_applied = 'last_name'
    direction_applied = 'asc'

    def get_queryset(self):
        user = self.request.user
        base_queryset = Lead.objects.exclude(status=Lead.StatusChoices.CONVERTED)
        queryset = self._filter_queryset_by_role(user, base_queryset)
        filter_params = self.request.GET.copy()
        filter_params.pop('sort', None)
        filter_params.pop('dir', None)
        filter_params.pop('page', None)
        self.filterset = LeadFilter(filter_params, queryset=queryset)
        queryset = self.filterset.qs

        sort_by_param = self.request.GET.get('sort', 'last_name')
        direction_param = self.request.GET.get('dir', 'asc')
        valid_sort_fields = [
            'last_name',
            'first_name',
            'company_name',
            'status',
            'source',
            'territory__name',
            'assigned_to__username',
            'updated_at',
        ]
        sort_by_validated = sort_by_param.lstrip('-')
        if sort_by_validated not in valid_sort_fields:
            sort_by_validated = 'last_name'
            direction_validated = 'asc'
        else:
            direction_validated = 'desc' if direction_param == 'desc' else 'asc'
        self.sort_by_applied = sort_by_validated
        self.direction_applied = direction_validated
        sort_by_final = (
            f'-{self.sort_by_applied}' if self.direction_applied == 'desc'
            else self.sort_by_applied
        )
        queryset = queryset.order_by(sort_by_final).select_related(
            'territory',
            'assigned_to',
        )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filterset'] = self.filterset
        context['sort_by'] = self.sort_by_applied
        context['direction'] = self.direction_applied
        context['opposite_direction'] = (
            'desc' if self.direction_applied == 'asc' else 'asc'
        )
        query_params = self.request.GET.copy()
        query_params.pop('sort', None)
        query_params.pop('dir', None)
        query_params.pop('page', None)
        context['current_filters_encoded'] = query_params.urlencode()
        return context


class LeadDetailView(BaseCrmView, DetailView):
    model = Lead
    context_object_name = 'lead'
    template_name = 'crm_entities/lead_detail.html'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related(
            'territory',
            'assigned_to',
            'created_by',
        )
        return self._filter_queryset_by_role(user, queryset)


class LeadCreateView(BaseCrmView, CreateView):
    model = Lead
    form_class = LeadForm
    template_name = 'crm_entities/lead_form.html'
    success_url = reverse_lazy('crm_entities:lead-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Create New Lead'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if hasattr(form.instance, 'assigned_to') and not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)


class LeadUpdateView(BaseCrmView, UpdateView):
    model = Lead
    form_class = LeadForm
    template_name = 'crm_entities/lead_form.html'
    success_url = reverse_lazy('crm_entities:lead-list')

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Update Lead: {self.object.full_name}'
        return context


class LeadDeleteView(BaseCrmView, DeleteView):
    model = Lead
    template_name = 'crm_entities/lead_confirm_delete.html'
    success_url = reverse_lazy('crm_entities:lead-list')
    context_object_name = 'lead'

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        return self._filter_queryset_by_role(user, queryset)


class LeadConvertView(BaseCrmView, View):
    """Lead conversion view with permission checks and transaction handling"""
    def post(self, request, *args, **kwargs):
        lead_pk = kwargs.get('pk')
        lead = Lead.objects.filter(pk=lead_pk).first()

        if not lead:
            messages.error(request, f"Lead with ID {lead_pk} not found.")
            return redirect('crm_entities:lead-list')

        user = request.user

        # Permission check
        has_permission = False
        if user.is_admin_role:
            has_permission = True
        elif user.is_manager_role:
            try:
                managed_territories = user.managed_territories.all()
                team_members = CustomUser.objects.filter(
                    territory__in=managed_territories,
                    role=CustomUser.Roles.SALES,
                ).exclude(pk=user.pk)
                if (
                    lead.assigned_to == user or
                    lead.created_by == user or
                    lead.assigned_to in team_members or
                    lead.created_by in team_members or
                    (lead.territory and lead.territory in managed_territories)
                ):
                    has_permission = True
            except Exception as e:
                print(f"Error checking manager permission for lead conversion: {e}")
                if lead.assigned_to == user or lead.created_by == user:
                    has_permission = True
        elif user.is_sales_role:
            if lead.assigned_to == user or lead.created_by == user:
                has_permission = True

        if not has_permission:
            messages.error(request, "You do not have permission to convert this lead.")
            return redirect('crm_entities:lead-detail', pk=lead.pk)

        # Status Checks
        if lead.status == Lead.StatusChoices.CONVERTED:
            messages.warning(request, f"Lead '{lead.full_name}' is already converted.")
            return redirect('crm_entities:lead-detail', pk=lead.pk)
        if lead.status == Lead.StatusChoices.LOST:
            messages.error(request, "Cannot convert a 'Lost' lead.")
            return redirect('crm_entities:lead-detail', pk=lead.pk)
        if lead.status != Lead.StatusChoices.QUALIFIED:
            messages.warning(
                request,
                f"Lead '{lead.full_name}' must be 'Qualified' to be converted."
            )
            return redirect('crm_entities:lead-detail', pk=lead.pk)

        # Conversion Logic
        try:
            with transaction.atomic():
                # Create Account
                account_name = (
                    lead.company_name or f"{lead.full_name}'s Company (from Lead)"
                )
                if Account.objects.filter(name=account_name).exists():
                    raise ValueError(
                        f"An account with the name '{account_name}' already exists. "
                        "Cannot convert lead automatically. Please resolve manually."
                    )
                account = Account.objects.create(
                    name=account_name,
                    phone_number=lead.work_phone,
                    billing_address=lead.address,
                    created_by=user,
                    assigned_to=lead.assigned_to or user,
                    territory=lead.territory,
                )
                messages.info(request, f"New Account created: '{account.name}'")

                # Create Contact
                contact = Contact.objects.create(
                    first_name=lead.first_name,
                    last_name=lead.last_name,
                    account=account,
                    title=lead.title,
                    department=lead.department,
                    email=lead.email,
                    work_phone=lead.work_phone,
                    mobile_phone_1=lead.mobile_phone_1,
                    mobile_phone_2=lead.mobile_phone_2,
                    created_by=user,
                    assigned_to=lead.assigned_to or user,
                    notes=f"Converted from Lead: {lead.full_name}\n" + (
                        lead.notes or ""
                    ),
                )
                messages.info(request, f"New Contact created: '{contact.full_name}'")

                # Create Deal
                default_close_date = timezone.now().date() + timedelta(days=30)
                deal = Deal.objects.create(
                    name=f"{account.name} - Opportunity from Lead {lead.last_name}",
                    account=account,
                    primary_contact=contact,
                    stage=Deal.StageChoices.QUALIFICATION,
                    amount=0.00,
                    currency='PHP',
                    close_date=default_close_date,
                    created_by=user,
                    assigned_to=lead.assigned_to or user,
                    description=(
                        f"Created from converted Lead: {lead.full_name}\n"
                        f"Amount needs review.\n" + (lead.notes or "")
                    ),
                )
                messages.info(request, f"New Deal created: '{deal.name}'")

                # Update Lead Status
                lead.status = Lead.StatusChoices.CONVERTED
                lead.save(update_fields=['status', 'updated_at'])

            messages.success(
                request,
                f"Lead '{lead.full_name}' converted successfully!"
            )
            return redirect('crm_entities:account-detail', pk=account.pk)

        except Exception as e:
            print(f"ERROR during Lead Conversion (PK: {lead_pk}): {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error converting lead: {str(e)}")
            return redirect('crm_entities:lead-detail', pk=lead.pk)


class AccountAutocomplete(LoginRequiredMixin, autocomplete.Select2QuerySetView):
    def get_queryset(self):
        user = self.request.user
        qs = Account.objects.all()
        if user.is_admin_role:
            pass
        elif user.is_manager_role:
            try:
                managed_territories = user.managed_territories.all()
                team_members = CustomUser.objects.filter(
                    territory__in=managed_territories,
                ).exclude(pk=user.pk)
                qs = qs.filter(
                    Q(assigned_to=user) |
                    Q(created_by=user) |
                    Q(assigned_to__in=team_members) |
                    Q(created_by__in=team_members) |
                    Q(territory__in=managed_territories)
                ).distinct()
            except Exception:
                qs = qs.filter(
                    Q(assigned_to=user) | Q(created_by=user)
                ).distinct()
        elif user.is_sales_role:
            qs = qs.filter(
                Q(assigned_to=user) | Q(created_by=user)
            ).distinct()
        else:
            qs = Account.objects.none()

        if self.q:
            qs = qs.filter(name__icontains=self.q)

        return qs.order_by('name')


class ContactAutocomplete(LoginRequiredMixin, autocomplete.Select2QuerySetView):
    def get_queryset(self):
        user = self.request.user
        if user.is_admin_role:
            qs = Contact.objects.all()
        else:
            qs = Contact.objects.all()

        account = None
        deal_pk = self.forwarded.get('deal', None)
        if deal_pk:
            try:
                deal = Deal.objects.select_related('account').get(pk=deal_pk)
                account = deal.account
            except Deal.DoesNotExist:
                return Contact.objects.none()
        if account:
            qs = qs.filter(account=account)

        if self.q:
            qs = qs.filter(
                Q(first_name__icontains=self.q) |
                Q(last_name__icontains=self.q) |
                Q(email__icontains=self.q)
            ).distinct()
        return qs.order_by('last_name', 'first_name')


class LeadAutocomplete(LoginRequiredMixin, autocomplete.Select2QuerySetView):
    def get_queryset(self):
        user = self.request.user
        qs = Lead.objects.exclude(
            status__in=[Lead.StatusChoices.CONVERTED, Lead.StatusChoices.LOST]
        )

        if user.is_admin_role:
            pass
        elif user.is_manager_role:
            try:
                managed_territories = user.managed_territories.all()
                team_members = CustomUser.objects.filter(
                    territory__in=managed_territories,
                ).exclude(pk=user.pk)
                qs = qs.filter(
                    Q(assigned_to=user) |
                    Q(created_by=user) |
                    Q(assigned_to__in=team_members) |
                    Q(created_by__in=team_members) |
                    Q(territory__in=managed_territories)
                ).distinct()
            except Exception:
                qs = qs.filter(
                    Q(assigned_to=user) | Q(created_by=user)
                ).distinct()
        elif user.is_sales_role:
            qs = qs.filter(
                Q(assigned_to=user) | Q(created_by=user)
            ).distinct()
        else:
            qs = Lead.objects.none()

        if self.q:
            qs = qs.filter(
                Q(first_name__icontains=self.q) |
                Q(last_name__icontains=self.q) |
                Q(company_name__icontains=self.q)
            ).distinct()
        return qs.order_by('last_name', 'first_name')


@login_required
def account_export_view(request):
    if not request.user.is_admin_role:
        return HttpResponseForbidden("Permission Denied.")

    base_queryset = Account.objects.all()
    filter_params = request.GET.copy()
    filter_params.pop('sort', None)
    filter_params.pop('dir', None)
    filter_params.pop('page', None)
    filterset = AccountFilter(filter_params, queryset=base_queryset)
    queryset = filterset.qs
    sort_by_param = request.GET.get('sort', 'name')
    direction_param = request.GET.get('dir', 'asc')
    valid_sort_fields = [
        'name',
        'status',
        'territory__name',
        'assigned_to__username',
        'updated_at',
    ]
    sort_by_validated = sort_by_param.lstrip('-')
    if sort_by_validated not in valid_sort_fields:
        sort_by_validated = 'name'
        direction_validated = 'asc'
    else:
        direction_validated = 'desc' if direction_param == 'desc' else 'asc'
    sort_by_final = (
        f'-{sort_by_validated}' if direction_validated == 'desc'
        else sort_by_validated
    )
    queryset = queryset.order_by(sort_by_final).select_related(
        'territory',
        'assigned_to',
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts"
    headers = [
        "ID",
        "Name",
        "Website",
        "Phone",
        "Billing Address",
        "Shipping Address",
        "Industry",
        "Status",
        "Territory",
        "Assigned To",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    for account in queryset:
        assigned_to_name = (
            account.assigned_to.get_full_name() or
            account.assigned_to.username if account.assigned_to else ""
        )
        created_at_formatted = (
            localtime(account.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if account.created_at else ""
        )
        updated_at_formatted = (
            localtime(account.updated_at).strftime('%Y-%m-%d %H:%M:%S')
            if account.updated_at else ""
        )
        row = [
            account.pk,
            account.name,
            account.website or "",
            account.phone_number or "",
            account.billing_address or "",
            account.shipping_address or "",
            account.industry or "",
            account.status or "",
            account.territory.name if account.territory else "",
            assigned_to_name,
            created_at_formatted,
            updated_at_formatted,
        ]
        ws.append(row)
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = 'attachment; filename="accounts_export.xlsx"'
    wb.save(response)
    return response


@login_required
def contact_export_view(request):
    if not request.user.is_admin_role:
        return HttpResponseForbidden("Permission Denied.")

    base_queryset = Contact.objects.all()
    filter_params = request.GET.copy()
    filter_params.pop('sort', None)
    filter_params.pop('dir', None)
    filter_params.pop('page', None)
    filterset = ContactFilter(filter_params, queryset=base_queryset)
    queryset = filterset.qs
    sort_by_param = request.GET.get('sort', 'last_name')
    direction_param = request.GET.get('dir', 'asc')
    valid_sort_fields = [
        'last_name',
        'first_name',
        'account__name',
        'title',
        'department',
        'email',
        'assigned_to__username',
        'updated_at',
    ]
    sort_by_validated = sort_by_param.lstrip('-')
    if sort_by_validated not in valid_sort_fields:
        sort_by_validated = 'last_name'
        direction_validated = 'asc'
    else:
        direction_validated = 'desc' if direction_param == 'desc' else 'asc'
    sort_by_final = (
        f'-{sort_by_validated}' if direction_validated == 'desc'
        else sort_by_validated
    )
    queryset = queryset.order_by(sort_by_final).select_related(
        'account',
        'assigned_to',
        'created_by',
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    headers = [
        "ID",
        "First Name",
        "Last Name",
        "Account",
        "Title",
        "Department",
        "Email",
        "Work Phone",
        "Mobile 1",
        "Mobile 2",
        "Notes",
        "Assigned To",
        "Created By",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    for contact in queryset:
        assigned_to_name = (
            contact.assigned_to.get_full_name() or
            contact.assigned_to.username if contact.assigned_to else ""
        )
        created_by_name = (
            contact.created_by.get_full_name() or
            contact.created_by.username if contact.created_by else ""
        )
        created_at_formatted = (
            localtime(contact.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if contact.created_at else ""
        )
        updated_at_formatted = (
            localtime(contact.updated_at).strftime('%Y-%m-%d %H:%M:%S')
            if contact.updated_at else ""
        )
        row = [
            contact.pk,
            contact.first_name or "",
            contact.last_name or "",
            contact.account.name if contact.account else "",
            contact.title or "",
            contact.department or "",
            contact.email or "",
            contact.work_phone or "",
            contact.mobile_phone_1 or "",
            contact.mobile_phone_2 or "",
            contact.notes or "",
            assigned_to_name,
            created_by_name,
            created_at_formatted,
            updated_at_formatted,
        ]
        ws.append(row)
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = 'attachment; filename="contacts_export.xlsx"'
    wb.save(response)
    return response


@login_required
def lead_export_view(request):
    if not request.user.is_admin_role:
        return HttpResponseForbidden("Permission Denied.")

    base_queryset = Lead.objects.all()
    filter_params = request.GET.copy()
    filter_params.pop('sort', None)
    filter_params.pop('dir', None)
    filter_params.pop('page', None)
    filterset = LeadFilter(filter_params, queryset=base_queryset)
    queryset = filterset.qs
    sort_by_param = request.GET.get('sort', 'last_name')
    direction_param = request.GET.get('dir', 'asc')
    valid_sort_fields = [
        'last_name',
        'first_name',
        'company_name',
        'status',
        'source',
        'territory__name',
        'assigned_to__username',
        'updated_at',
    ]
    sort_by_validated = sort_by_param.lstrip('-')
    if sort_by_validated not in valid_sort_fields:
        sort_by_validated = 'last_name'
        direction_validated = 'asc'
    else:
        direction_validated = 'desc' if direction_param == 'desc' else 'asc'
    sort_by_final = (
        f'-{sort_by_validated}' if direction_validated == 'desc'
        else sort_by_validated
    )
    queryset = queryset.order_by(sort_by_final).select_related(
        'territory',
        'assigned_to',
        'created_by',
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = [
        "ID",
        "First Name",
        "Last Name",
        "Company",
        "Title",
        "Department",
        "Email",
        "Work Phone",
        "Mobile 1",
        "Mobile 2",
        "Address",
        "Notes",
        "Status",
        "Source",
        "Territory",
        "Assigned To",
        "Created By",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)
    for lead in queryset:
        assigned_to_name = (
            lead.assigned_to.get_full_name() or
            lead.assigned_to.username if lead.assigned_to else ""
        )
        created_by_name = (
            lead.created_by.get_full_name() or
            lead.created_by.username if lead.created_by else ""
        )
        created_at_formatted = (
            localtime(lead.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if lead.created_at else ""
        )
        updated_at_formatted = (
            localtime(lead.updated_at).strftime('%Y-%m-%d %H:%M:%S')
            if lead.updated_at else ""
        )
        row = [
            lead.pk,
            lead.first_name or "",
            lead.last_name or "",
            lead.company_name or "",
            lead.title or "",
            lead.department or "",
            lead.email or "",
            lead.work_phone or "",
            lead.mobile_phone_1 or "",
            lead.mobile_phone_2 or "",
            lead.address or "",
            lead.notes or "",
            lead.get_status_display(),
            lead.get_source_display() or "",
            lead.territory.name if lead.territory else "",
            assigned_to_name,
            created_by_name,
            created_at_formatted,
            updated_at_formatted,
        ]
        ws.append(row)
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = 'attachment; filename="leads_export.xlsx"'
    wb.save(response)
    return response